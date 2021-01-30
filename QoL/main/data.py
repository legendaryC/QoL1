from .models import *
import datetime
import openpyxl
from pathlib import Path
import random
import re


class ImportData:
    def __init__(self):
        super().__init__()

    def getID(self, s):
        res = '0000'+str(s)
        return 'pa'+res[-4:]

    def get_data(self):
        # new_entry = Albumin(patient_ID='pa0001',
        #                     date_time=datetime.datetime(2020, 5, 17), value=1.2)  kljlk
        # new_entry.save()

        xlsx_file = Path('/home/chen7874/QoL1/QoL/main', 'dataSet.xlsx')
        wb_obj = openpyxl.load_workbook(xlsx_file)

        # Read the active sheet:
        sheet = wb_obj.active
        # print(sheet["C1"].value)
        i = None
        for row in sheet.iter_rows(max_row=sheet.max_row-1):
            # print(row[0].value, end='/n')
            if row[0].value:
                i = row[0].value
                if type(row[3].value) is str:
                    info_day, info_month, info_year = re.split(
                        '-|/', row[3].value)
                else:
                    info_day, info_month, info_year = row[3].value.day, row[3].value.month, row[3].value.year
                if Medical_Info.objects.filter(patient_ID=self.getID(i)):
                    continue
                medical_Info = Medical_Info(patient_ID=self.getID(i),
                                            date_time=datetime.datetime(int(info_year), int(info_month), int(info_day)), ckd=float(row[16].value), first_dialysis_date=datetime.datetime(int(info_year), int(info_month), int(info_day)), number_of_dialysis=int(row[4].value), expected_number_of_dialysis=int(row[25].value), dialysis_frequency=2)

                medical_Info.save()
                # new_entry.save()

                # for cell in row:
                #     print(cell.value, end=" ")
                # print()
            #NOTE: Lab_data
            print("##", str(row[6].value))
            if type(row[6].value) is str:
                lab_day, lab_month, lab_year = re.split('-|/', row[6].value)
            else:
                lab_day, lab_month, lab_year = row[6].value.day, row[6].value.month, row[6].value.year

            if Albumin.objects.filter(patient_ID=self.getID(i), date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day))):
                continue
            albumin = Albumin(patient_ID=self.getID(i),
                              date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=float(row[7].value))
            albumin.save()
            bicarbonate = Bicarbonate(patient_ID=self.getID(i),
                                      date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=float(row[8].value))
            bicarbonate.save()
            bUN = BUN(patient_ID=self.getID(i),
                      date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=float(row[9].value))
            bUN.save()
            calcium = Calcium(patient_ID=self.getID(i),
                              date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=float(row[10].value))
            calcium.save()
            creatinine = Creatinine(patient_ID=self.getID(i),
                                    date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=float(row[11].value))
            creatinine.save()
            hemoglobin = Hemoglobin(patient_ID=self.getID(i),
                                    date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=float(row[12].value))
            hemoglobin.save()
            potassium = Potassium(patient_ID=self.getID(i),
                                  date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=float(row[13].value))
            potassium.save()
            phosphorus = Phosphorus(patient_ID=self.getID(i),
                                    date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=float(row[14].value))
            phosphorus.save()
            sodium = Sodium(patient_ID=self.getID(i),
                            date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=float(row[15].value))
            sodium.save()

            pTH = PTH(patient_ID=self.getID(i),
                      date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=random.randint(0, 70))
            pTH.save()
            alkaline_Phosphatase = Alkaline_Phosphatase(patient_ID=self.getID(i),
                                                        date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), value=random.randint(0, 300))
            alkaline_Phosphatase.save()

            #NOTE: comorbidities
            comorbidities = Comorbidities(patient_ID=self.getID(i),
                                          date_time=datetime.datetime(int(lab_year), int(lab_month), int(lab_day)), hypertension=int(row[17].value), diabetes=int(row[18].value), cardiovascular_disease=random.randint(0, 1), typhoid=random.randint(0, 1))
            comorbidities.save()

            #NOTE: Dialysis_data
            if type(row[20].value) is str:

                dialysis_day, dialysis_month, dialysis_year = re.split(
                    '-|/', row[20].value)
            else:
                dialysis_day, dialysis_month, dialysis_year = row[
                    20].value.day, row[20].value.month, row[20].value.year

            # weight = random.randint(80, 180)
            dialysis = Dialysis(patient_ID=self.getID(i), date_time=datetime.datetime(int(dialysis_year), int(dialysis_month), int(
                dialysis_day)), bp=str(random.randint(100, 140))+"/"+str(random.randint(60, 80)), weight=float(row[22].value), kt_v_ratio=float(row[23].value), temperature=random.randint(950, 1020)/10, pulse_rate=random.randint(60, 100), dialysis_duration=4)
            dialysis.save()
            # NOTE: Before_Dialysis:
            # if int(dialysis_day) == 0:
            #     before_dialysis = Dialysis(patient_ID=self.getID(i), date_time=datetime.datetime(int(dialysis_year), int(dialysis_month)-1, 30), bp=random.randint(100, 140)+"/"+random.randint(
            #         60, 80), weight=weight, kt_v_ratio=float(row[23].value), temperature=random.randint(950, 1020)/10, pulse_rate=random.randint(60, 100), dialysis_duration=4)

            #     before_dialysis.save()
            # else:
            #     before_dialysis = Dialysis(patient_ID=self.getID(i), date_time=datetime.datetime(int(dialysis_year), int(dialysis_month), int(dialysis_day)-1), bp=str(random.randint(
            #         100, 140))+"/"+str(random.randint(60, 80)), weight=weight, kt_v_ratio=float(row[23].value), temperature=random.randint(950, 1020)/10, pulse_rate=random.randint(60, 100), dialysis_duration=4)
            #     before_dialysis.save()

            # else:
            #     for cell in row:
            #         print(i, cell.value, end=" ")
            #     print()
